### DEBUG test file for Excel export with synchronized calendar and original formatting
### Can delete - if Excel-only script now works 

##################################################
# Final Synchronized Version: Fixed Calendar + Original Formatting
##################################################

import snowflake.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import sys

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

def get_connection():
    return snowflake.connector.connect(
        user=os.environ["SNOWFLAKE_USER"],
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
        role=os.environ.get("SNOWFLAKE_ROLE"),
        authenticator=os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    )

def query_snowflake_activity(sql):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    conn.close()
    df = pd.DataFrame(rows, columns=columns)
    if not df.empty:
        df["ACTIVITY_DATE"] = pd.to_datetime(df["ACTIVITY_DATE"])
    return df

ECDS_PROVIDERS = [
    "University College London Hospitals NHS Foundation Trust",
    "Whittington Health NHS Trust",
    "Royal Free London NHS Foundation Trust",
    "Moorfields Eye Hospital NHS Foundation Trust",
]

def calculate_dynamic_summary(df_apc, df_op, df_ecds, all_providers_list, ecds_providers, calendar_dates):
    """Calculates missing days by comparing data against the Master Calendar."""
    summary_data = []
    
    for provider in all_providers_list:
        p_apc = df_apc[df_apc["PROVIDER"] == provider]
        p_op = df_op[df_op["PROVIDER"] == provider]
        p_ecds = df_ecds[df_ecds["PROVIDER"] == provider]
        
        # Check against the actual dates in the Master Calendar
        apc_missing = sum(1 for d in calendar_dates if d not in p_apc["ACTIVITY_DATE"].dt.floor('D').values)
        op_missing = sum(1 for d in calendar_dates if d not in p_op["ACTIVITY_DATE"].dt.floor('D').values)
        
        if provider in ecds_providers:
            ecds_missing = sum(1 for d in calendar_dates if d not in p_ecds["ACTIVITY_DATE"].dt.floor('D').values)
        else:
            ecds_missing = 0
            
        total = apc_missing + op_missing + ecds_missing
        summary_data.append({
            "PROVIDER": provider,
            "APC_MISSING_DAYS": apc_missing,
            "OP_MISSING_DAYS": op_missing,
            "ECDS_MISSING_DAYS": ecds_missing,
            "TOTAL_MISSING_SUBMISSIONS": total,
            "ACTION_REQUIRED": "Contact ISL" if total > 0 else "All Complete"
        })
    return pd.DataFrame(summary_data)

def build_pivot_table(ws, df, title, start_row, calendar_dates):
    """Build pivot table with full formatting and anomaly detection."""
    day_labels = [d.strftime("%d/%m/%Y") for d in calendar_dates]
    weekday_labels = [d.strftime("%a") for d in calendar_dates]
    providers = sorted(df["PROVIDER"].unique()) if not df.empty else ["No Data Found"]

    # Calculate Stats for Anomaly Detection (using weekday vs weekend logic)
    provider_stats = {}
    if not df.empty:
        for p in providers:
            p_data = df[df["PROVIDER"] == p]
            wd = p_data[p_data["ACTIVITY_DATE"].dt.weekday < 5]["RECORDS"]
            we = p_data[p_data["ACTIVITY_DATE"].dt.weekday >= 5]["RECORDS"]
            provider_stats[p] = {
                "weekday": {"mean": wd.mean(), "std": wd.std()} if len(wd[wd>0]) > 2 else None,
                "weekend": {"mean": we.mean(), "std": we.std()} if len(we[we>0]) > 2 else None
            }

    # Pivot and Reindex to match Master Calendar exactly
    if not df.empty:
        df_labeled = df.copy()
        df_labeled["DAY_LABEL"] = df_labeled["ACTIVITY_DATE"].dt.strftime("%d/%m/%Y")
        pivot = df_labeled.pivot(index="PROVIDER", columns="DAY_LABEL", values="RECORDS")
        pivot = pivot.reindex(columns=day_labels)
    else:
        pivot = pd.DataFrame(index=providers, columns=day_labels)

    # Drawing Headers
    timestamp = datetime.now().strftime("dbt Pipeline run at %H:%M GMT, %d %b %Y")
    ws.cell(row=start_row, column=1, value=timestamp).font = Font(italic=True)
    ws.cell(row=start_row+2, column=1, value=title).font = Font(bold=True, size=12)
    header_row = start_row + 3

    # Weekday Row
    for i, day_name in enumerate(weekday_labels, start=2):
        cell = ws.cell(row=header_row, column=i, value=day_name)
        cell.alignment = Alignment(horizontal="center")
        if day_name in ['Sat', 'Sun']:
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Date Row
    ws.cell(row=header_row+1, column=1, value="Provider").font = Font(bold=True)
    for i, d_label in enumerate(day_labels, start=2):
        cell = ws.cell(row=header_row+1, column=i, value=d_label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for provider in providers:
        row_values = [provider]
        for d_label in day_labels:
            val = pivot.loc[provider, d_label]
            row_values.append("MISSING" if pd.isna(val) or val == 0 else int(val))
        ws.append(row_values)

    # Formatting and Anomaly Colors
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+2, max_row=ws.max_row, min_col=1, max_col=len(day_labels)+1)):
        p_name = providers[row_idx]
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if col_idx == 1: continue # Skip provider name
            
            if cell.value == "MISSING":
                cell.fill = red_fill
            else:
                # Anomaly Logic
                dt_obj = calendar_dates[col_idx-2]
                is_we = dt_obj.weekday() >= 5
                stats = provider_stats.get(p_name, {}).get("weekend" if is_we else "weekday")
                
                if stats and stats["std"] and stats["std"] > 0:
                    z = abs((float(cell.value) - stats["mean"]) / stats["std"])
                    cell.fill = orange_fill if z > 3 else (yellow_fill if z > 2 else green_fill)
                else:
                    cell.fill = green_fill

    ws.column_dimensions["A"].width = 35
    for i in range(2, len(day_labels)+2):
        ws.column_dimensions[get_column_letter(i)].width = 11.5

def build_summary_table(ws, df, title, start_row, is_unstable):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    offset = 1
    if is_unstable:
        warn = ws.cell(row=start_row+1, column=1, value="⚠️ Warning: Last 14 days data may still be updating")
        warn.font = Font(color="FF0000", italic=True)
        offset = 2
    
    headers = ["Provider", "APC Missing", "OP Missing", "ECDS Missing", "Total Missing", "Action Required"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row+offset, column=i, value=h)
        c.font = Font(bold=True)
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for _, row in df.iterrows():
        ws.append(list(row.values))
        last_row = ws.max_row
        fill = PatternFill(start_color="FFC7CE" if row["TOTAL_MISSING_SUBMISSIONS"] > 0 else "C6EFCE", fill_type="solid")
        for i in range(1, 7):
            ws.cell(row=last_row, column=i).fill = fill
            ws.cell(row=last_row, column=i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def export_to_excel(stable_bundle, unstable_bundle, filename="provider_status.xlsx"):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Stable Data (60 Days)"
    build_summary_table(ws1, stable_bundle["summary"], "Provider Missing Days Summary (60 Days)", 1, False)
    build_pivot_table(ws1, stable_bundle["apc"], "Inpatient Daily Status", ws1.max_row+2, stable_bundle["calendar"])
    build_pivot_table(ws1, stable_bundle["op"], "Outpatient Daily Status", ws1.max_row+3, stable_bundle["calendar"])
    build_pivot_table(ws1, stable_bundle["ecds"], "ECDS Daily Status", ws1.max_row+3, stable_bundle["calendar"])
    ws1.freeze_panes = "B2"

    ws2 = wb.create_sheet("Unstable Data (Last 14 Days)")
    build_summary_table(ws2, unstable_bundle["summary"], "Provider Missing Days Summary (Last 14 Days)", 1, True)
    build_pivot_table(ws2, unstable_bundle["apc"], "Inpatient Daily Status (Unstable)", ws2.max_row+2, unstable_bundle["calendar"])
    build_pivot_table(ws2, unstable_bundle["op"], "Outpatient Daily Status (Unstable)", ws2.max_row+3, unstable_bundle["calendar"])
    build_pivot_table(ws2, unstable_bundle["ecds"], "ECDS Daily Status (Unstable)", ws2.max_row+3, unstable_bundle["calendar"])
    ws2.freeze_panes = "B2"
    
    wb.save(filename)
    os.startfile(filename)
    return filename

if __name__ == "__main__":
    # 1. Setup Master Calendars
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    unstable_cal = pd.date_range(end=today - timedelta(days=1), periods=14).tolist()
    stable_cal = pd.date_range(end=today - timedelta(days=15), periods=60).tolist()
    
    # 2. Query Data
    sql_unstable = f"WHERE ACTIVITY_DATE >= '{unstable_cal[0].strftime('%Y-%m-%d')}'"
    sql_stable = f"WHERE ACTIVITY_DATE >= '{stable_cal[0].strftime('%Y-%m-%d')}' AND ACTIVITY_DATE <= '{stable_cal[-1].strftime('%Y-%m-%d')}'"
    
    print("Querying Snowflake...")
    df_apc_u = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_APC_ACTIVITY_DBT {sql_unstable}")
    df_op_u = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_OP_ACTIVITY_DBT {sql_unstable}")
    df_ecds_u = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_ECDS_ACTIVITY_DBT {sql_unstable}")
    
    df_apc_s = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_APC_ACTIVITY_DBT {sql_stable}")
    df_op_s = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_OP_ACTIVITY_DBT {sql_stable}")
    df_ecds_s = query_snowflake_activity(f"SELECT PROVIDER, ACTIVITY_DATE, RECORDS FROM PROVIDER_DAILY_ECDS_ACTIVITY_DBT {sql_stable}")

    # 3. Master Provider List
    all_providers = sorted(set(df_apc_s["PROVIDER"].unique()) | set(df_op_s["PROVIDER"].unique()))
    
    # 4. Summaries & Export
    summ_u = calculate_dynamic_summary(df_apc_u, df_op_u, df_ecds_u, all_providers, ECDS_PROVIDERS, unstable_cal)
    summ_s = calculate_dynamic_summary(df_apc_s, df_op_s, df_ecds_s, all_providers, ECDS_PROVIDERS, stable_cal)

    export_to_excel(
        {"summary": summ_s, "apc": df_apc_s, "op": df_op_s, "ecds": df_ecds_s, "calendar": stable_cal},
        {"summary": summ_u, "apc": df_apc_u, "op": df_op_u, "ecds": df_ecds_u, "calendar": unstable_cal}
    )
    print("Done! Report saved and opened.")