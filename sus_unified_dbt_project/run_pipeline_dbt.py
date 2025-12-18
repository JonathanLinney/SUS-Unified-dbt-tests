##################################################
# Script to run dbt models/tests, query them in Snowflake, and then export the results to Excel
##################################################
# Script runs from the sus_unified_dbt_project folder - set by 'project_dir' variable
##################################################

## Set the number of days to report on in MAIN PIPELINE RUN SECTION at the end of this script ##

import subprocess
import snowflake.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

def run_dbt():
    print("Running dbt models...")
    project_dir = "c:/dev/dbt test/sus_unified_dbt_project"

    subprocess.run(["dbt", "run"], check=True, cwd=project_dir)

    print("Running dbt tests...")
    try:
        subprocess.run(["dbt", "test"], check=True, cwd=project_dir)
    except subprocess.CalledProcessError:
        print("⚠️ dbt test failed, continuing with pipeline...")

def query_snowflake_activity(sql):
    conn = snowflake.connector.connect(
        user=os.environ["SNOWFLAKE_USER"],
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
        role=os.environ.get("SNOWFLAKE_ROLE"),
        authenticator=os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    )
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    df = pd.DataFrame(rows, columns=columns)
    df["ACTIVITY_DATE"] = pd.to_datetime(df["ACTIVITY_DATE"])
    df["DAY_LABEL"] = df["ACTIVITY_DATE"].dt.strftime("%d/%m/%Y")
    return df

def query_snowflake_summary():
    conn = snowflake.connector.connect(
        user=os.environ["SNOWFLAKE_USER"],
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
        role=os.environ.get("SNOWFLAKE_ROLE"),
        authenticator=os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            PROVIDER, 
            APC_MISSING_DAYS,
            OP_MISSING_DAYS,
            ECDS_MISSING_DAYS,
            TOTAL_MISSING_SUBMISSIONS,
            ACTION_REQUIRED
        FROM PROVIDER_MISSING_SUMMARY
    """)
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    df = pd.DataFrame(rows, columns=columns)
    return df

def build_summary_table(ws, df, title, start_row):
    # Title row
    ws.cell(row=start_row, column=1, value=title).alignment = Alignment(horizontal="left")

    # Header row
    headers = ["Provider", "APC Missing", "OP Missing", "ECDS Missing", "Total Missing", "Action Required"]
    ws.append(headers)

    # Data rows
    for _, row in df.iterrows():
        ws.append([
            row["PROVIDER"], 
            row["APC_MISSING_DAYS"], 
            row["OP_MISSING_DAYS"],
            row["ECDS_MISSING_DAYS"],
            row["TOTAL_MISSING_SUBMISSIONS"], 
            row["ACTION_REQUIRED"]
        ])

    # Formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    header_row = start_row + 1
    first_data_row = header_row + 1
    last_row = ws.max_row
    last_col = 6  # Provider, APC, OP, ECDS, Total, Action

    # Header borders
    for cell in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=last_col):
        for c in cell:
            c.border = thin_border

    # Data borders + fills
    for row in ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=1, max_col=last_col):
        provider, apc, op, ecds, total, action = row
        for c in row:
            c.border = thin_border
        
        if int(total.value) > 0:
            for c in row:
                c.fill = red_fill
        else:
            for c in row:
                c.fill = green_fill

def build_pivot_table(ws, df, title, start_row):
    df["ACTIVITY_DATE"] = pd.to_datetime(df["ACTIVITY_DATE"])
    df["DAY_LABEL"] = df["ACTIVITY_DATE"].dt.strftime("%d/%m/%Y")
    df["WEEKDAY"] = df["ACTIVITY_DATE"].dt.day_name().str[:3]  # Mon, Tue, ...

    day_order = df[["ACTIVITY_DATE", "DAY_LABEL", "WEEKDAY"]].drop_duplicates().sort_values("ACTIVITY_DATE")
    day_labels_sorted = day_order["DAY_LABEL"].tolist()
    weekday_labels_sorted = day_order["WEEKDAY"].tolist()
    providers = sorted(df["PROVIDER"].unique())

    pivot_records = df.pivot(index="PROVIDER", columns="DAY_LABEL", values="RECORDS")
    pivot_records = pivot_records.reindex(index=providers, columns=day_labels_sorted)

    # Calculate weekday/weekend stats per provider
    provider_stats = {}
    for provider in providers:
        provider_data = df[df["PROVIDER"] == provider]
        weekday_data = provider_data[provider_data["ACTIVITY_DATE"].dt.weekday < 5]["RECORDS"]
        weekend_data = provider_data[provider_data["ACTIVITY_DATE"].dt.weekday >= 5]["RECORDS"]

        stats = {}
        for label, series in [("weekday", weekday_data), ("weekend", weekend_data)]:
            valid = series[series.notna() & (series > 0)]
            if len(valid) > 2:
                stats[label] = {"mean": valid.mean(), "std": valid.std()}
            else:
                stats[label] = None
        provider_stats[provider] = stats

    def map_status(val):
        if pd.isna(val):
            return "MISSING"
        try:
            return "MISSING" if float(val) == 0 else int(val)
        except Exception:
            return "MISSING"

    pivot = pivot_records.map(map_status)

    timestamp = datetime.now().strftime("dbt Pipeline run at %H:%M GMT, %d %b %Y")
    ws.cell(row=start_row, column=1, value=timestamp).alignment = Alignment(horizontal="left")
    ws.cell(row=start_row+1, column=1, value="")
    ws.cell(row=start_row+2, column=1, value=title)

    header_row = start_row+3

    # Weekday row above date headers
    ws.append([""] + weekday_labels_sorted)
    for idx, cell in enumerate(ws[header_row], start=1):
        cell.alignment = Alignment(horizontal="center")
        if idx > 1:  # skip Provider column
            activity_date = day_order.iloc[idx-2]["ACTIVITY_DATE"]
            if activity_date.weekday() >= 5:  # Saturday/Sunday
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                     top=Side(style="medium"), bottom=Side(style="medium"))

    # Date header row
    ws.append(["Provider"] + day_labels_sorted)
    for idx, cell in enumerate(ws[header_row+1], start=1):
        cell.alignment = Alignment(horizontal="center")
        if idx > 1:
            activity_date = day_order.iloc[idx-2]["ACTIVITY_DATE"]
            if activity_date.weekday() >= 5:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                     top=Side(style="medium"), bottom=Side(style="medium"))

    # Data rows
    for provider, row in pivot.iterrows():
        ws.append([provider] + list(row.values))

    # Formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    first_data_row = header_row+2
    last_row = ws.max_row
    last_col = len(pivot.columns) + 1

    # Borders for headers
    for row in ws.iter_rows(min_row=header_row, max_row=header_row+1, min_col=1, max_col=last_col):
        for c in row:
            c.border = thin_border

    # Data rows with weekend-aware anomaly detection
    for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=1, max_col=last_col)):
        provider_name = providers[row_idx]
        for j, c in enumerate(row, start=1):
            c.border = thin_border
            if j == 1:  # Provider column
                continue
            if c.value == "MISSING":
                c.fill = red_fill
            elif isinstance(c.value, (int, float)):
                # Determine if this column is weekend or weekday
                activity_date = day_order.iloc[j-2]["ACTIVITY_DATE"]
                is_weekend = activity_date.weekday() >= 5
                stats = provider_stats[provider_name]["weekend" if is_weekend else "weekday"]

                if stats and stats["std"] > 0:
                    z_score = abs((float(c.value) - stats["mean"]) / stats["std"])
                    if z_score > 3:
                        c.fill = orange_fill
                    elif z_score > 2:
                        c.fill = yellow_fill
                    else:
                        c.fill = green_fill
                else:
                    c.fill = green_fill
            else:
                c.fill = green_fill

    # Column widths
    ws.column_dimensions["A"].width = 35
    for i in range(2, last_col+1):
        ws.column_dimensions[get_column_letter(i)].width = 11.36


def export_to_excel(df_summary, df_inpatient, df_op, df_ecds, filename="provider_status.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Provider Daily Status"

    # Summary block
    build_summary_table(ws, df_summary, "Provider Missing Days Summary (Rolling 20 Day Monitoring Window)", start_row=1)

    # Spacer
    ws.append([]); ws.append([])

    # Inpatient block
    build_pivot_table(ws, df_inpatient, "Inpatient Provider Daily Status", start_row=ws.max_row+1)

    # Outpatient block
    build_pivot_table(ws, df_op, "Outpatient Provider Daily Status", start_row=ws.max_row+3)

    # ECDS block
    build_pivot_table(ws, df_ecds, "Emergency Attendances (ECDS) Daily Status", start_row=ws.max_row+3)

    ws.freeze_panes = ws["B2"] # Freeze top row and first column
    wb.save(filename)
    print(f"Excel report saved as {filename}")
    return filename

def open_excel(filename):
    try:
        os.startfile(filename)
    except Exception:
        pass

if __name__ == "__main__":
    run_dbt()
    df_summary   = query_snowflake_summary()
    df_inpatient = query_snowflake_activity("""
    SELECT PROVIDER, ACTIVITY_DATE, RECORDS
    FROM PROVIDER_DAILY_APC_ACTIVITY_DBT
    WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '34 days'
    AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
""")

df_op = query_snowflake_activity("""
    SELECT PROVIDER, ACTIVITY_DATE, RECORDS
    FROM PROVIDER_DAILY_OP_ACTIVITY_DBT
    WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '34 days'
    AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
""")

df_ecds = query_snowflake_activity("""
    SELECT PROVIDER, ACTIVITY_DATE, RECORDS
    FROM PROVIDER_DAILY_ECDS_ACTIVITY_DBT
    WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '34 days'
    AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
""")

    if df_summary.empty or df_inpatient.empty or df_op.empty or df_ecds.empty:
        print("One or more datasets are empty. Please check dbt models and Snowflake sources.")
        sys.exit(0)

    filename = export_to_excel(df_summary, df_inpatient, df_op, df_ecds)
    open_excel(filename)
    print("dbt Project Pipeline completed successfully!")