##################################################
# Script to run query the dbt models in Snowflake (but not actually run dbt), and export the results to Excel
##################################################

import snowflake.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

def get_connection():
    """Create and return a Snowflake connection."""
    return snowflake.connector.connect(
        user=os.environ["SNOWFLAKE_USER"],
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
        role=os.environ.get("SNOWFLAKE_ROLE"),
        authenticator=os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    )

def query_snowflake_activity(sql):
    """Query activity data and return as DataFrame."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    conn.close()
   
    df = pd.DataFrame(rows, columns=columns)
    df["ACTIVITY_DATE"] = pd.to_datetime(df["ACTIVITY_DATE"])
    df["DAY_LABEL"] = df["ACTIVITY_DATE"].dt.strftime("%d/%m/%Y")
    return df

# Trusts that are expected to submit ECDS (ie. have A&E departments)
ECDS_PROVIDERS = [
    "University College London Hospitals NHS Foundation Trust",
    "Whittington Health NHS Trust",
    "Royal Free London NHS Foundation Trust",
    "Moorfields Eye Hospital NHS Foundation Trust",
]

def calculate_dynamic_summary(df_apc, df_op, df_ecds, all_providers_list=None, ecds_providers=None): 
    """ 
    Calculate summary of missing days dynamically from the filtered datasets. 
    Allows restricting ECDS to only providers that actually have ECDS feeds. 
    """

    # Default: if no ECDS list provided, assume all providers have ECDS    
    # - If explicitly provided, use it 
    # - Else, if we have a master provider list, assume all have ECDS (backwards compatible) 
    # - Else, no ECDS providers 
    if ecds_providers is None: 
        if all_providers_list is not None: 
            ecds_providers = list(all_providers_list) 
        else: 
            ecds_providers = []

    # Helper: densify a dataset with optional max date forcing
    def densify(df, force_max_date=None):
        if df.empty:
            return df
        df = df.copy()
        
        # Use forced max date if provided, otherwise use data's max
        min_date = df["ACTIVITY_DATE"].min()
        max_date = pd.to_datetime(force_max_date) if force_max_date is not None else df["ACTIVITY_DATE"].max()
        
        all_dates = pd.date_range(min_date, max_date, freq='D')
        providers = df["PROVIDER"].unique()

        full_grid = pd.MultiIndex.from_product(
            [providers, all_dates],
            names=["PROVIDER", "ACTIVITY_DATE"]
        ).to_frame(index=False)

        merged = full_grid.merge(df[["PROVIDER", "ACTIVITY_DATE", "RECORDS"]],
                                 on=["PROVIDER", "ACTIVITY_DATE"],
                                 how="left")
        merged["RECORDS"] = merged["RECORDS"].fillna(0)
        return merged

    # Store the maximum dates from each dataset BEFORE filling missing providers
    apc_max_date = df_apc["ACTIVITY_DATE"].max() if not df_apc.empty else None
    op_max_date = df_op["ACTIVITY_DATE"].max() if not df_op.empty else None
    ecds_max_date = df_ecds["ACTIVITY_DATE"].max() if not df_ecds.empty else None

    # If we have a master provider list, fill missing providers appropriately
    if all_providers_list:
        # Determine date range across all datasets
        all_dates = pd.concat([
            df_apc["ACTIVITY_DATE"] if not df_apc.empty else pd.Series(dtype='datetime64[ns]'),
            df_op["ACTIVITY_DATE"] if not df_op.empty else pd.Series(dtype='datetime64[ns]'),
            df_ecds["ACTIVITY_DATE"] if not df_ecds.empty else pd.Series(dtype='datetime64[ns]')
        ])

        if len(all_dates) > 0:
            date_range = pd.date_range(start=all_dates.min(), end=all_dates.max(), freq='D')

            for provider in all_providers_list:

                # APC fill
                if provider not in df_apc["PROVIDER"].values:
                    df_apc = pd.concat([df_apc, pd.DataFrame({
                        "PROVIDER": [provider] * len(date_range),
                        "ACTIVITY_DATE": date_range,
                        "RECORDS": [0] * len(date_range)
                    })], ignore_index=True)

                # OP fill
                if provider not in df_op["PROVIDER"].values:
                    df_op = pd.concat([df_op, pd.DataFrame({
                        "PROVIDER": [provider] * len(date_range),
                        "ACTIVITY_DATE": date_range,
                        "RECORDS": [0] * len(date_range)
                    })], ignore_index=True)

                # ECDS fill — **only for providers that actually have ECDS**
                if provider in ecds_providers and provider not in df_ecds["PROVIDER"].values:
                    df_ecds = pd.concat([df_ecds, pd.DataFrame({
                        "PROVIDER": [provider] * len(date_range),
                        "ACTIVITY_DATE": date_range,
                        "RECORDS": [0] * len(date_range)
                    })], ignore_index=True)

    # Densify - use the max dates we stored earlier to ensure all providers extend to the same endpoint
    apc_full = densify(df_apc, force_max_date=apc_max_date)
    op_full = densify(df_op, force_max_date=op_max_date)
    ecds_full = densify(df_ecds, force_max_date=ecds_max_date)

    # Build provider list for summary
    all_providers = sorted(all_providers_list) if all_providers_list else sorted(
        set(df_apc["PROVIDER"]) | set(df_op["PROVIDER"]) | set(df_ecds["PROVIDER"])
    )

    summary_data = []
    for provider in all_providers:
        apc_missing = len(apc_full[(apc_full["PROVIDER"] == provider) & (apc_full["RECORDS"] == 0)])
        op_missing = len(op_full[(op_full["PROVIDER"] == provider) & (op_full["RECORDS"] == 0)])

        # ECDS only counts if provider is expected to have ECDS
        if provider in ecds_providers:
            ecds_missing = len(ecds_full[(ecds_full["PROVIDER"] == provider) & (ecds_full["RECORDS"] == 0)])
        else:
            ecds_missing = 0

        total = apc_missing + op_missing + ecds_missing

        summary_data.append({
            "PROVIDER": provider,
            "APC_MISSING_DAYS": apc_missing,
            "OP_MISSING_DAYS": op_missing,
            "ECDS_MISSING_DAYS": ecds_missing,
            "TOTAL_MISSING_SUBMISSIONS": total,
            "ACTION_REQUIRED": "Contact ISL about missing submissions" if total > 0 else "All submissions complete"
        })

    return pd.DataFrame(summary_data), df_apc, df_op, df_ecds

def build_summary_table(ws, df, title, start_row, is_unstable=False):
    """Build summary table in Excel worksheet."""
    # Title row with warning for unstable data
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.alignment = Alignment(horizontal="left")
    title_cell.font = Font(bold=True, size=12)
   
    if is_unstable:
        warning_cell = ws.cell(row=start_row+1, column=1,
                              value="⚠️ Warning: This data covers the last 14 days and may still be updating")
        warning_cell.font = Font(color="FF0000", italic=True)
        warning_cell.alignment = Alignment(horizontal="left")
        header_row_offset = 2
    else:
        header_row_offset = 1

    # Header row
    headers = ["Provider", "APC Missing", "OP Missing", "ECDS Missing", "Total Missing", "Action Required"]
    header_row = start_row + header_row_offset
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for _, row in df.iterrows():
        ws.append([
            row["PROVIDER"],
            row["APC_MISSING_DAYS"],
            row["OP_MISSING_DAYS"],
            row["ECDS_MISSING_DAYS"],
            row["TOTAL_MISSING_SUBMISSIONS"],
            row["ACTION_REQUIRED"]
        ])

    # Formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    first_data_row = header_row + 1
    last_row = ws.max_row
    last_col = 6  # Provider, APC, OP, ECDS, Total, Action

    # Header borders
    for cell in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=last_col):
        for c in cell:
            c.border = thin_border

    # Data borders + fills
    for row in ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=1, max_col=last_col):
        provider, apc, op, ecds, total, action = row
        for c in row:
            c.border = thin_border
       
        if int(total.value) > 0:
            for c in row:
                c.fill = red_fill
        else:
            for c in row:
                c.fill = green_fill

def build_pivot_table(ws, df, title, start_row):
    """Build pivot table with daily activity in Excel worksheet."""
    # Handle empty dataframe
    if df.empty:
        ws.cell(row=start_row, column=1, value=f"{title} - No data available")
        return
    
    df["ACTIVITY_DATE"] = pd.to_datetime(df["ACTIVITY_DATE"])
    df["DAY_LABEL"] = df["ACTIVITY_DATE"].dt.strftime("%d/%m/%Y")
    df["WEEKDAY"] = df["ACTIVITY_DATE"].dt.day_name().str[:3]  # Mon, Tue, ...

    day_order = df[["ACTIVITY_DATE", "DAY_LABEL", "WEEKDAY"]].drop_duplicates().sort_values("ACTIVITY_DATE")
    
    # If day_order is empty, we can't build a pivot table
    if day_order.empty:
        ws.cell(row=start_row, column=1, value=f"{title} - No dates available")
        return
    
    day_labels_sorted = day_order["DAY_LABEL"].tolist()
    weekday_labels_sorted = day_order["WEEKDAY"].tolist()
    providers = sorted(df["PROVIDER"].unique())

    # If no providers, return
    if not providers:
        ws.cell(row=start_row, column=1, value=f"{title} - No providers available")
        return

    pivot_records = df.pivot(index="PROVIDER", columns="DAY_LABEL", values="RECORDS")
    pivot_records = pivot_records.reindex(index=providers, columns=day_labels_sorted)

    # Calculate weekday/weekend stats per provider
    provider_stats = {}
    for provider in providers:
        provider_data = df[df["PROVIDER"] == provider]
        weekday_data = provider_data[provider_data["ACTIVITY_DATE"].dt.weekday < 5]["RECORDS"]
        weekend_data = provider_data[provider_data["ACTIVITY_DATE"].dt.weekday >= 5]["RECORDS"]

        stats = {}
        for label, series in [("weekday", weekday_data), ("weekend", weekend_data)]:
            valid = series[series.notna() & (series > 0)]
            if len(valid) > 2:
                stats[label] = {"mean": valid.mean(), "std": valid.std()}
            else:
                stats[label] = None
        provider_stats[provider] = stats

    def map_status(val):
        if pd.isna(val):
            return "MISSING"
        try:
            return "MISSING" if float(val) == 0 else int(val)
        except Exception:
            return "MISSING"

    pivot = pivot_records.map(map_status)

    timestamp = datetime.now().strftime("dbt Pipeline run at %H:%M GMT, %d %b %Y")
    ws.cell(row=start_row, column=1, value=timestamp).alignment = Alignment(horizontal="left")
    ws.cell(row=start_row+1, column=1, value="")
    ws.cell(row=start_row+2, column=1, value=title)

    header_row = start_row+3

    # Weekday row above date headers
    ws.append([""] + weekday_labels_sorted)
    for idx, cell in enumerate(ws[header_row], start=1):
        cell.alignment = Alignment(horizontal="center")
        if idx > 1:  # skip Provider column
            # BOUNDS CHECK: Make sure idx-2 is within day_order range
            day_idx = idx - 2
            if day_idx < len(day_order):
                activity_date = day_order.iloc[day_idx]["ACTIVITY_DATE"]
                if activity_date.weekday() >= 5:  # Saturday/Sunday
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                         top=Side(style="medium"), bottom=Side(style="medium"))

    # Date header row
    ws.append(["Provider"] + day_labels_sorted)
    for idx, cell in enumerate(ws[header_row+1], start=1):
        cell.alignment = Alignment(horizontal="center")
        if idx > 1:
            # BOUNDS CHECK: Make sure idx-2 is within day_order range
            day_idx = idx - 2
            if day_idx < len(day_order):
                activity_date = day_order.iloc[day_idx]["ACTIVITY_DATE"]
                if activity_date.weekday() >= 5:
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                                         top=Side(style="medium"), bottom=Side(style="medium"))

    # Data rows
    for provider, row in pivot.iterrows():
        ws.append([provider] + list(row.values))

    # Formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    first_data_row = header_row+2
    last_row = ws.max_row
    last_col = len(pivot.columns) + 1

    # Borders for headers
    for row in ws.iter_rows(min_row=header_row, max_row=header_row+1, min_col=1, max_col=last_col):
        for c in row:
            c.border = thin_border

    # Data rows with weekend-aware anomaly detection
    for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=1, max_col=last_col)):
        provider_name = providers[row_idx]
        for j, c in enumerate(row, start=1):
            c.border = thin_border
            if j == 1:  # Provider column
                continue
            if c.value == "MISSING":
                c.fill = red_fill
            elif isinstance(c.value, (int, float)):
                # BOUNDS CHECK: Determine if this column is weekend or weekday
                day_idx = j - 2
                if day_idx < len(day_order):
                    activity_date = day_order.iloc[day_idx]["ACTIVITY_DATE"]
                    is_weekend = activity_date.weekday() >= 5
                    stats = provider_stats[provider_name]["weekend" if is_weekend else "weekday"]

                    if stats and stats["std"] > 0:
                        z_score = abs((float(c.value) - stats["mean"]) / stats["std"])
                        if z_score > 3:
                            c.fill = orange_fill
                        elif z_score > 2:
                            c.fill = yellow_fill
                        else:
                            c.fill = green_fill
                    else:
                        c.fill = green_fill
                else:
                    # Fallback if out of bounds
                    c.fill = green_fill
            else:
                c.fill = green_fill

    # Column widths
    ws.column_dimensions["A"].width = 35
    for i in range(2, last_col+1):
        ws.column_dimensions[get_column_letter(i)].width = 11.36


def export_to_excel(stable_data, unstable_data, filename="provider_status.xlsx"):
    """
    Export data to Excel with two sheets:
    1. Stable data (60 days, excluding last 14)
    2. Unstable data (last 14 days)
    """
    wb = Workbook()
   
    # ===== STABLE DATA SHEET (60 days) =====
    ws_stable = wb.active
    ws_stable.title = "Stable Data (60 Days)"
   
    # Summary block
    build_summary_table(
        ws_stable,
        stable_data["summary"],
        "Provider Missing Days Summary (60 Days Stable Data, Excluding Last 14 Days)",
        start_row=1,
        is_unstable=False
    )
   
    # Spacer
    ws_stable.append([]); ws_stable.append([])
   
    # Inpatient block
    build_pivot_table(ws_stable, stable_data["apc"], "Inpatient Provider Daily Status", start_row=ws_stable.max_row+1)
   
    # Outpatient block
    build_pivot_table(ws_stable, stable_data["op"], "Outpatient Provider Daily Status", start_row=ws_stable.max_row+3)
   
    # ECDS block
    build_pivot_table(ws_stable, stable_data["ecds"], "Emergency Attendances (ECDS) Daily Status", start_row=ws_stable.max_row+3)
   
    ws_stable.freeze_panes = ws_stable["B2"]
   
    # ===== UNSTABLE DATA SHEET (14 days) =====
    ws_unstable = wb.create_sheet(title="Unstable Data (Last 14 Days)")
   
    # Summary block with warning
    build_summary_table(
        ws_unstable,
        unstable_data["summary"],
        "Provider Missing Days Summary (Last 14 Days - Data Still Updating)",
        start_row=1,
        is_unstable=True
    )
   
    # Spacer
    ws_unstable.append([]); ws_unstable.append([])
   
    # Inpatient block
    build_pivot_table(ws_unstable, unstable_data["apc"], "Inpatient Provider Daily Status (Unstable)", start_row=ws_unstable.max_row+1)
   
    # Outpatient block
    build_pivot_table(ws_unstable, unstable_data["op"], "Outpatient Provider Daily Status (Unstable)", start_row=ws_unstable.max_row+3)
   
    # ECDS block
    build_pivot_table(ws_unstable, unstable_data["ecds"], "Emergency Attendances (ECDS) Daily Status (Unstable)", start_row=ws_unstable.max_row+3)
   
    ws_unstable.freeze_panes = ws_unstable["B2"]
   
    wb.save(filename)
    print(f"Excel report saved as {filename}")
    return filename

def open_excel(filename):
    """Attempt to open Excel file."""
    try:
        os.startfile(filename)
    except Exception:
        pass

if __name__ == "__main__":
    print("="*60)
    print("Excel Report Builder from dbt Snowflake Data")
    print("="*60)

   
    print("\n" + "="*60)
    print("Querying Snowflake for stable data (60 days)...")
    print("="*60)
   
    # ===== STABLE DATA (60 days, excluding last 14) =====
    df_apc_stable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_APC_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '74 days'
        AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
    """)
   
    df_op_stable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_OP_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '74 days'
        AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
    """)
   
    df_ecds_stable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_ECDS_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '74 days'
        AND ACTIVITY_DATE < CURRENT_DATE - INTERVAL '14 days'
    """)
   
    # Get master list of all providers from stable data
    all_providers = sorted(set(
        list(df_apc_stable["PROVIDER"].unique()) +
        list(df_op_stable["PROVIDER"].unique()) +
        list(df_ecds_stable["PROVIDER"].unique())
    ))
   
    print(f"Found {len(all_providers)} providers: {', '.join(all_providers)}")
   
    # Calculate dynamic summary for stable data with full provider list
    # PASS the master list and the ECDS specific list to ensure 0-record rows are generated
    df_summary_stable, df_apc_stable, df_op_stable, df_ecds_stable = calculate_dynamic_summary(
    df_apc_stable, 
    df_op_stable, 
    df_ecds_stable,
    all_providers_list=all_providers, # Ensure we use the full list
    ecds_providers=ECDS_PROVIDERS     # Ensure ECDS-only Provider subset is respected
)
   
    print("\n" + "="*60)
    print("Querying Snowflake for unstable data (last 14 days)...")
    print("="*60)
   
    # ===== UNSTABLE DATA (last 14 days) =====
    df_apc_unstable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_APC_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '14 days'
    """)
   
    df_op_unstable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_OP_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '14 days'
    """)
   
    df_ecds_unstable = query_snowflake_activity("""
        SELECT PROVIDER, ACTIVITY_DATE, RECORDS
        FROM PROVIDER_DAILY_ECDS_ACTIVITY_DBT
        WHERE ACTIVITY_DATE >= CURRENT_DATE - INTERVAL '14 days'
    """)
   
    # Calculate dynamic summary for unstable data with full provider list (plus ECDS-only filtering)
    # This ensures providers with NO submissions in last 14 days still appear
    df_summary_unstable, df_apc_unstable, df_op_unstable, df_ecds_unstable = calculate_dynamic_summary(
    df_apc_unstable,
    df_op_unstable,
    df_ecds_unstable,
    all_providers_list=all_providers,
    ecds_providers=ECDS_PROVIDERS
)

   
    # Check for empty datasets
    if df_apc_stable.empty or df_op_stable.empty or df_ecds_stable.empty:
        print("⚠️  One or more stable datasets are empty. Please check dbt models and Snowflake sources.")
        sys.exit(1)
   
    print("\n" + "="*60)
    print("Generating Excel report...")
    print("="*60)
   
    # Package data for export
    stable_data = {
        "summary": df_summary_stable,
        "apc": df_apc_stable,
        "op": df_op_stable,
        "ecds": df_ecds_stable
    }
   
    unstable_data = {
        "summary": df_summary_unstable,
        "apc": df_apc_unstable,
        "op": df_op_unstable,
        "ecds": df_ecds_unstable
    }
   
    filename = export_to_excel(stable_data, unstable_data)
    open_excel(filename)
   
    print("\n" + "="*60)
    print("✅ Excel Summary Report completed successfully!")
    print("="*60)
    print(f"📊 Stable data: 60 days (excluding last 14)")
    print(f"⚠️ Unstable data: Last 14 days")
    print(f"👥 Providers tracked: {len(all_providers)}")
    print(f"📁 File: {filename}")
    print("="*60)